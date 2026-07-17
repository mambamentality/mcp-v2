# tools/extract_attendance.py
"""Parsea la matriz de asistencia (personas en columnas, fecha por fila)."""
from __future__ import annotations

import base64
import io
import json
import logging
from typing import Any, Dict, List, Optional

import azure.functions as func
from openpyxl import load_workbook

from .mcp_helpers import ToolProperty, get_arguments, tool_properties_json

logger = logging.getLogger(__name__)

_TOOL_PROPERTIES = tool_properties_json(
    [
        ToolProperty("fileBase64", "string", "Excel de lista de asistencia en base64.", isRequired=True),
        ToolProperty("fecha", "string", "Fecha de la sesión a extraer (opcional).", isRequired=False),
    ]
)

_HEADER_ANCHOR = "Fecha Programada"


def register_extract_attendance_tool(app: func.FunctionApp):
    @app.mcp_tool_trigger(
        arg_name="context",
        tool_name="ExtractAttendance",
        description="Extrae asistentes, horario y asistencia desde el Excel de lista de asistencia.",
        tool_properties=_TOOL_PROPERTIES,
    )
    def extract_attendance(context) -> str:
        try:
            args = get_arguments(context)
            file_b64 = args.get("fileBase64")
            if not file_b64:
                raise ValueError("Debe enviar 'fileBase64'.")

            wb = load_workbook(io_bytesio(base64.b64decode(file_b64)), data_only=True)
            sheet = wb.active

            header_row_idx, name_cols, total_col, hora_inicio_col, hora_fin_col = _find_header(sheet)
            target_fecha = args.get("fecha")

            data_row = _find_data_row(sheet, header_row_idx, target_fecha)
            if data_row is None:
                raise ValueError("No se encontró una fila de sesión que coincida.")

            attendees = []
            for col_idx, name in name_cols.items():
                value = sheet.cell(row=data_row, column=col_idx).value
                attendees.append({"nombre": name, "asistio": bool(value)})

            hora_inicio = sheet.cell(row=data_row, column=hora_inicio_col).value if hora_inicio_col else None
            hora_fin = sheet.cell(row=data_row, column=hora_fin_col).value if hora_fin_col else None

            return json.dumps(
                {
                    "status": "ok",
                    "asistentes": attendees,
                    "horaInicio": str(hora_inicio) if hora_inicio else None,
                    "horaFin": str(hora_fin) if hora_fin else None,
                },
                ensure_ascii=False,
            )
        except Exception as exc:
            logger.exception("ExtractAttendance error")
            return json.dumps({"status": "error", "message": str(exc)}, ensure_ascii=False)

    return extract_attendance


def io_bytesio(data: bytes):
    import io as _io
    return _io.BytesIO(data)


def _find_header(sheet):
    for row in sheet.iter_rows(min_row=1, max_row=10):
        for cell in row:
            if cell.value and _HEADER_ANCHOR.lower() in str(cell.value).lower():
                return _map_columns(sheet, cell.row)
    raise ValueError(f"No se encontró el encabezado '{_HEADER_ANCHOR}' en las primeras 10 filas.")


def _map_columns(sheet, header_row_idx: int):
    name_cols: Dict[int, str] = {}
    total_col: Optional[int] = None
    hora_inicio_col: Optional[int] = None
    hora_fin_col: Optional[int] = None

    capturing = False
    for cell in sheet[header_row_idx]:
        value = str(cell.value).strip() if cell.value else ""
        if value.lower() == "tipo":
            capturing = True
            continue
        if value == "Total":
            total_col = cell.column
            capturing = False
            continue
        if "hora real de inicio" in value.lower():
            hora_inicio_col = cell.column
            continue
        if "hora real de finalizaci" in value.lower():
            hora_fin_col = cell.column
            continue
        if capturing and value:
            name_cols[cell.column] = value

    return header_row_idx, name_cols, total_col, hora_inicio_col, hora_fin_col

# tools/extract_attendance.py — reemplazar _find_data_row

from .date_utils import normalize_date


def _find_data_row(sheet, header_row_idx: int, target_fecha: Optional[str]) -> Optional[int]:
    target_normalized = normalize_date(target_fecha) if target_fecha else None

    for row in sheet.iter_rows(min_row=header_row_idx + 1, max_row=header_row_idx + 200):
        first_cell = row[0]
        if first_cell.value is None:
            continue
        if target_normalized is None:
            return first_cell.row
        if normalize_date(first_cell.value) == target_normalized:
            return first_cell.row
    return None