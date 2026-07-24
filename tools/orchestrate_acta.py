# tools/orchestrate_acta.py — VERSIÓN CONSOLIDADA COMPLETA, reemplaza el archivo entero
import json
import logging
import re
from typing import Any, Dict, List
from uuid import uuid4

import azure.functions as func

from . import sharepoint_client, paths, maestro_store, people_catalog
from .acta_draft_store import ActaDraftStore
from .extract_convocatoria import _extract_text as _pdf_text, _first_match, _detect_modalidad, _parse_orden_del_dia, _parse_roles, _DATE_RE, _TIME_RE, _PLACE_RE
from .extract_attendance import _find_header, _find_data_row, io_bytesio
from .extract_backup_content import _extract_pdf, _extract_docx, _extract_pptx, _extract_xlsx, _find_cite
from .match_backup_to_agenda import extract_point_number
from .maestro_matching import match_to_maestro
from .narrative_filler import fill_narratives
from .intro_paragraph_builder import build_intro_fields
from .mcp_helpers import ToolProperty, get_arguments, tool_properties_json

logger = logging.getLogger(__name__)

_TOOL_PROPERTIES = tool_properties_json(
    [
        ToolProperty("gestion", "string", "Año de gestión (ej. '2026').", isRequired=True),
        ToolProperty("directorio", "string", "Nombre del directorio/fecha de sesión (ej. '2026-07-28').", isRequired=True),
        ToolProperty("numeroActa", "string", "Número de acta (ej. '08/2026'). Si se omite, se sugiere uno.", isRequired=False),
    ]
)

_EXCLUDE_KEYWORDS = ["convocatoria", "asistencia", "lista de asistencia", "acta maestra", "catalogo_personas"]


def register_process_gestion_directory_tool(app: func.FunctionApp):
    store = ActaDraftStore()

    @app.mcp_tool_trigger(
        arg_name="context",
        tool_name="ProcessGestionDirectory",
        description="Corre todo el proceso de extracción (convocatoria, asistencia, respaldos, cruce con Acta Maestra) y deja un borrador de acta listo para revisión.",
        tool_properties=_TOOL_PROPERTIES,
    )
    def process_gestion_directory(context) -> str:
        try:
            args = get_arguments(context)
            gestion = args.get("gestion")
            directorio = args.get("directorio")
            if not gestion or not directorio:
                raise ValueError("Debe enviar 'gestion' y 'directorio'.")

            folder_path = paths.insumos_folder(gestion, directorio)
            items = sharepoint_client.list_children(folder_path)
            file_map = {item["name"]: item for item in items if not item["isFolder"]}

            convocatoria_data = _process_convocatoria(file_map, folder_path)
            attendance_data = _process_attendance(file_map, folder_path, convocatoria_data)

            maestro_rows = maestro_store.read_maestro()
            agenda_matched = match_to_maestro(convocatoria_data["ordenDelDia"], maestro_rows)
            agenda_with_backups = _attach_backups(agenda_matched, file_map, folder_path)

            fecha_corte = convocatoria_data.get("fecha", "")
            fill_narratives(agenda_with_backups, fecha_corte=fecha_corte)

            presidenta = next((a["nombre"] for a in attendance_data["asistentes"] if a.get("rol") == "Presidencia"), "")

            acta_id = str(uuid4())
            numero_sugerido = args.get("numeroActa") or _suggest_numero_acta(gestion)

            acta_data: Dict[str, Any] = {
                "actaId": acta_id,
                "numeroActa": numero_sugerido,
                "numeroActaConfirmado": bool(args.get("numeroActa")),
                "fecha": convocatoria_data.get("fecha", ""),
                "fechaTexto": convocatoria_data.get("fecha", ""),  # alias: la plantilla usa este nombre
                "horaInicio": convocatoria_data.get("hora", ""),
                "horaFin": attendance_data.get("horaFin"),
                "lugar": convocatoria_data.get("lugar", ""),
                "modalidadSesion": convocatoria_data.get("modalidad", ""),
                "ordenDelDia": agenda_with_backups,
                "asistentes": attendance_data.get("asistentes", []),
                "roles": convocatoria_data.get("roles", {}),
                "status": "pendingModalidad",
                "sourceFolder": folder_path,
                "outputFolder": paths.generadas_folder(gestion, directorio),
                "gestion": gestion,
                "directorio": directorio,
                "presidenta": presidenta,
                "secretaria": "",  # se completa manualmente vía UpdateActaPoint antes de aprobar
            }
            acta_data.update(build_intro_fields(acta_data))
            store.save_acta(acta_id, acta_data)

            return json.dumps(
                {
                    "status": "ok",
                    "actaId": acta_id,
                    "resumen": {
                        "puntos": len(agenda_with_backups),
                        "asistentes": len(attendance_data.get("asistentes", [])),
                        "asistentesSinRolResuelto": sum(1 for a in attendance_data.get("asistentes", []) if not a.get("rol")),
                        "puntosNuevos": _count_estado(agenda_with_backups, "punto_nuevo"),
                        "puntosPorConfirmar": _count_estado(agenda_with_backups, "requiere_confirmacion"),
                    },
                },
                ensure_ascii=False,
            )
        except Exception as exc:
            logger.exception("ProcessGestionDirectory error")
            return json.dumps({"status": "error", "message": str(exc)}, ensure_ascii=False)

    return process_gestion_directory


def _process_convocatoria(file_map: Dict[str, Any], folder_path: str) -> Dict[str, Any]:
    conv_file = next((name for name in file_map if "convocatoria" in name.lower()), None)
    if not conv_file:
        raise ValueError("No se encontró el archivo de Convocatoria en el directorio.")

    content = sharepoint_client.download_file(f"{folder_path}/{conv_file}")
    text = _pdf_text(content)

    return {
        "fecha": _first_match(_DATE_RE, text),
        "hora": _first_match(_TIME_RE, text),
        "lugar": _first_match(_PLACE_RE, text),
        "modalidad": _detect_modalidad(text),
        "ordenDelDia": _parse_orden_del_dia(text),
        "roles": _parse_roles(text),
    }


def _process_attendance(file_map: Dict[str, Any], folder_path: str, convocatoria_data: Dict[str, Any]) -> Dict[str, Any]:
    att_file = next(
        (name for name in file_map if "asistencia" in name.lower() and name.lower().endswith((".xlsx", ".xlsm"))),
        None,
    )
    if not att_file:
        raise ValueError("No se encontró el Excel de Lista de Asistencia en el directorio.")

    content = sharepoint_client.download_file(f"{folder_path}/{att_file}")
    from openpyxl import load_workbook
    wb = load_workbook(io_bytesio(content), data_only=True)
    sheet = wb.active

    header_row_idx, name_cols, total_col, hora_inicio_col, hora_fin_col = _find_header(sheet)

    data_row = _find_data_row(sheet, header_row_idx, convocatoria_data.get("fecha"))
    if data_row is None:
        data_row = _find_data_row(sheet, header_row_idx, None)
    if data_row is None:
        raise ValueError("No se encontró una fila de sesión en el Excel de asistencia.")

    # Resolver cada nombre crudo del Excel contra el catálogo de personas,
    # para obtener rol/género/nombre canónico (necesario para los párrafos
    # de Alta Gerencia/Administración, que la Convocatoria no lista).
    catalog = people_catalog.read_catalog()

    attendees = []
    for col_idx, raw_name in name_cols.items():
        value = sheet.cell(row=data_row, column=col_idx).value
        resolved = people_catalog.resolve_name(raw_name, catalog)
        attendees.append({
            "nombre": resolved["nombreCompleto"] if resolved else raw_name,
            "asistio": bool(value),
            "rol": resolved.get("rol", "") if resolved else "",
            "genero": resolved.get("genero", "") if resolved else "",
            "cargo": resolved.get("cargo", "") if resolved else "",
            "modalidad": "",
        })

    return {
        "asistentes": attendees,
        "horaInicio": str(sheet.cell(row=data_row, column=hora_inicio_col).value) if hora_inicio_col else None,
        "horaFin": str(sheet.cell(row=data_row, column=hora_fin_col).value) if hora_fin_col else None,
    }


def _attach_backups(points: List[Dict[str, Any]], file_map: Dict[str, Any], folder_path: str) -> List[Dict[str, Any]]:
    backup_matches = {}
    for name in file_map:
        if any(k in name.lower() for k in _EXCLUDE_KEYWORDS):
            continue
        numero = extract_point_number(name)
        if numero:
            backup_matches[name] = numero

    def process(node: Dict[str, Any]) -> Dict[str, Any]:
        files_for_point = [name for name, numero in backup_matches.items() if numero == node["numero"]]
        texts, cite = [], None
        for file_name in files_for_point:
            content = sharepoint_client.download_file(f"{folder_path}/{file_name}")
            extension = file_name.lower().rsplit(".", 1)[-1]
            if extension == "pdf":
                text = _extract_pdf(content)
            elif extension == "docx":
                text = _extract_docx(content)
            elif extension == "pptx":
                text = _extract_pptx(content)
            elif extension in ("xlsx", "xlsm"):
                text = _extract_xlsx(content)
            else:
                continue
            texts.append(text)
            if cite is None:
                cite = _find_cite(text)

        node["textoRespaldo"] = "\n\n".join(texts)
        node["cite"] = cite
        node["subpuntos"] = [process(s) for s in node.get("subpuntos", [])]
        return node

    return [process(dict(p)) for p in points]


def _suggest_numero_acta(gestion: str) -> str:
    try:
        items = sharepoint_client.list_children(paths.generadas_root_for_gestion(gestion))
        actas_existentes = sum(1 for item in items if item["isFolder"])
        return f"{actas_existentes + 1:02d}/{gestion}"
    except Exception:
        return ""


def _count_estado(points: List[Dict[str, Any]], estado: str) -> int:
    total = 0
    for p in points:
        if p.get("maestroEstado") == estado:
            total += 1
        total += _count_estado(p.get("subpuntos", []), estado)
    return total