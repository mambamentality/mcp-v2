# tools/maestro_store.py — Acta Maestra en Excel, leída/escrita directo por Graph API (sin Power Automate)
from __future__ import annotations

import io
from typing import Any, Dict, List

from openpyxl import load_workbook

from . import sharepoint_client, paths

_COLUMNS = [
    "maestroId", "titulo", "gerencia", "periodicidad", "tipo", "plantillaFrase",
    "narrativaAnterior", "determinacionAnterior", "ultimoCite", "ultimaFecha", "vecesTratado",
]


def read_maestro(maestro_path: str | None = None) -> List[Dict[str, Any]]:
    path = maestro_path or paths.maestro_path()
    content = sharepoint_client.download_file(path)
    wb = load_workbook(io.BytesIO(content), data_only=True)
    sheet = wb["Acta Maestra"] if "Acta Maestra" in wb.sheetnames else wb.active

    header = [cell.value for cell in sheet[1]]
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        rows.append(dict(zip(header, row)))
    return rows


def upsert_maestro_rows(updates: List[Dict[str, Any]], maestro_path: str | None = None) -> None:
    path = maestro_path or paths.maestro_path()
    content = sharepoint_client.download_file(path)
    wb = load_workbook(io.BytesIO(content))
    sheet = wb["Acta Maestra"] if "Acta Maestra" in wb.sheetnames else wb.active

    header = [cell.value for cell in sheet[1]]
    id_col = header.index("maestroId") + 1

    existing_ids = {}
    for row in sheet.iter_rows(min_row=2):
        cell = row[id_col - 1]
        if cell.value:
            existing_ids[cell.value] = cell.row

    for update in updates:
        maestro_id = update.get("maestroId")
        if not maestro_id:
            continue
        if maestro_id in existing_ids:
            row_idx = existing_ids[maestro_id]
            for col_name, value in update.items():
                if col_name in header:
                    col_idx = header.index(col_name) + 1
                    sheet.cell(row=row_idx, column=col_idx, value=value)
            veces_col = header.index("vecesTratado") + 1
            actual = sheet.cell(row=row_idx, column=veces_col).value or 0
            sheet.cell(row=row_idx, column=veces_col, value=actual + 1)
        else:
            new_row = [update.get(col, "") for col in header]
            if "vecesTratado" in header:
                new_row[header.index("vecesTratado")] = 1
            sheet.append(new_row)

    buffer = io.BytesIO()
    wb.save(buffer)
    folder, file_name = path.rsplit("/", 1)
    sharepoint_client.upload_file(folder, file_name, buffer.getvalue())