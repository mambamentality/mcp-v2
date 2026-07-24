# tools/people_catalog.py
"""Catálogo de personas en Excel sobre SharePoint, leído/escrito directo
por Graph API — mismo patrón que maestro_store.py."""
from __future__ import annotations

import io
import re
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

from . import sharepoint_client

_CATALOG_PATH = "Actas/Catalogo_Personas.xlsx"
_COLUMNS = ["nombreCompleto", "alias", "rol", "genero", "cargo"]


def read_catalog(catalog_path: str = _CATALOG_PATH) -> List[Dict[str, Any]]:
    content = sharepoint_client.download_file(catalog_path)
    wb = load_workbook(io.BytesIO(content), data_only=True)
    sheet = wb.active

    header = [cell.value for cell in sheet[1]]
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        rows.append(dict(zip(header, row)))
    return rows


def upsert_person(nombre_completo: str, alias: List[str], rol: str, genero: str, cargo: str = "", catalog_path: str = _CATALOG_PATH) -> None:
    content = sharepoint_client.download_file(catalog_path)
    wb = load_workbook(io.BytesIO(content))
    sheet = wb.active

    header = [cell.value for cell in sheet[1]]
    name_col = header.index("nombreCompleto") + 1

    existing_row = None
    for row in sheet.iter_rows(min_row=2):
        if row[name_col - 1].value == nombre_completo:
            existing_row = row[0].row
            break

    values = {
        "nombreCompleto": nombre_completo,
        "alias": "|".join(alias),
        "rol": rol,
        "genero": genero,
        "cargo": cargo,
    }

    if existing_row:
        for col_name, value in values.items():
            col_idx = header.index(col_name) + 1
            sheet.cell(row=existing_row, column=col_idx, value=value)
    else:
        sheet.append([values.get(col, "") for col in header])

    buffer = io.BytesIO()
    wb.save(buffer)
    folder, file_name = catalog_path.rsplit("/", 1)
    sharepoint_client.upload_file(folder, file_name, buffer.getvalue())


_STRIP_AREA_RE = re.compile(r",\s*[A-Z]{2,6}$")


def _normalize(name: str) -> str:
    name = _STRIP_AREA_RE.sub("", name or "")
    return re.sub(r"\s+", " ", name.strip().lower())


def resolve_name(raw_name: str, catalog: List[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    normalized_input = _normalize(raw_name)

    for person in catalog:
        candidates = [person.get("nombreCompleto", "")] + (
            person.get("alias", "").split("|") if person.get("alias") else []
        )
        for candidate in candidates:
            if _normalize(candidate) == normalized_input:
                return person

    from difflib import SequenceMatcher
    best, best_score = None, 0.0
    for person in catalog:
        score = SequenceMatcher(None, normalized_input, _normalize(person.get("nombreCompleto", ""))).ratio()
        if score > best_score:
            best, best_score = person, score
    return best if best_score >= 0.85 else None