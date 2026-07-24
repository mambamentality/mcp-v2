"""
import os, json

with open("local.settings.json", encoding="utf-8") as f:
    settings = json.load(f)
    for key, value in settings["Values"].items():
        os.environ[key] = value

from tools import sharepoint_client
from tools.extract_convocatoria import (
    _extract_text, _first_match, _detect_modalidad,
    _parse_orden_del_dia, _parse_roles, _DATE_RE, _TIME_RE, _PLACE_RE,
)

content = sharepoint_client.download_file("Actas/Insumos/2024/23-04-2024/00. 2024.04.23 Convocatoria Directorio.pdf")
text = _extract_text(content)

print("FECHA:", _first_match(_DATE_RE, text))
print("HORA:", _first_match(_TIME_RE, text))
print("LUGAR:", _first_match(_PLACE_RE, text))
print("MODALIDAD:", _detect_modalidad(text))
print()
print("=== ORDEN DEL DIA ===")
orden = _parse_orden_del_dia(text)
print(f"Cantidad de puntos top-level: {len(orden)}")
print(json.dumps(orden, ensure_ascii=False, indent=2))
print()
print("=== ROLES ===")
print(json.dumps(_parse_roles(text), ensure_ascii=False, indent=2))
"""
"""
import os, json

with open("local.settings.json", encoding="utf-8") as f:
    settings = json.load(f)
    for key, value in settings["Values"].items():
        os.environ[key] = value

from tools import sharepoint_client
from tools.extract_attendance import _find_header, _find_data_row, io_bytesio
from openpyxl import load_workbook

content = sharepoint_client.download_file("Actas/Insumos/2024/23-04-2024/Lista_de_Asistencia.xlsx")
wb = load_workbook(io_bytesio(content), data_only=True)
sheet = wb.active

print("Nombre de la hoja activa:", sheet.title)
print("Dimensiones:", sheet.dimensions)

try:
    header_row_idx, name_cols, total_col, hora_inicio_col, hora_fin_col = _find_header(sheet)
    print(f"\nFila de encabezado: {header_row_idx}")
    print(f"Columnas de nombres encontradas: {name_cols}")
    print(f"Columna 'Total': {total_col}")
    print(f"Columna 'Hora real de inicio': {hora_inicio_col}")
    print(f"Columna 'Hora real de finalización': {hora_fin_col}")

    data_row = _find_data_row(sheet, header_row_idx, None)
    print(f"\nFila de datos encontrada: {data_row}")

    if data_row:
        print("\nAsistencia:")
        for col_idx, name in name_cols.items():
            value = sheet.cell(row=data_row, column=col_idx).value
            print(f"  {name}: {value}")
        if hora_inicio_col:
            print(f"\nHora inicio: {sheet.cell(row=data_row, column=hora_inicio_col).value}")
        if hora_fin_col:
            print(f"Hora fin: {sheet.cell(row=data_row, column=hora_fin_col).value}")
except Exception as exc:
    print(f"ERROR: {exc}")

"""
"""
# test_maestro.py
import os, json
with open("local.settings.json", encoding="utf-8") as f:
    for k, v in json.load(f)["Values"].items():
        os.environ[k] = v

from tools import maestro_store

rows = maestro_store.read_maestro()
print(f"Filas en Acta Maestra: {len(rows)}")
if rows:
    print("Columnas encontradas:", list(rows[0].keys()))
    print("Primera fila:", rows[0])
    """




"""
import os, json

with open("local.settings.json", encoding="utf-8") as f:
    settings = json.load(f)
    for key, value in settings["Values"].items():
        os.environ[key] = value

from tools import sharepoint_client
from tools.extract_backup_content import _extract_pdf, _find_cite

# probamos con un PDF de respaldo real, no escaneado (debería ser texto digital)
content = sharepoint_client.download_file(
    "Actas/Insumos/2024/23-04-2024/03.1 Informe de Inversiones al 31.03.2024.pdf"
)
text = _extract_pdf(content)

print("Largo del texto:", len(text))
print("Cite encontrado:", _find_cite(text))
print()
print("Primeros 1000 caracteres:")
print(text[:1000])

"""



"""   PRUEBA DE EXTRACCIÓN DE TEXTO DE LA CONVOCATORIA DE DIRECTORIO (PDF)
import os, json

with open("local.settings.json", encoding="utf-8") as f:
    settings = json.load(f)
    for key, value in settings["Values"].items():
        os.environ[key] = value

from tools import sharepoint_client
from tools.extract_convocatoria import (
    _extract_text, _first_match, _detect_modalidad,
    _parse_orden_del_dia, _parse_roles, _DATE_RE, _TIME_RE, _PLACE_RE,
)

content = sharepoint_client.download_file(
    "Actas/Insumos/2024/23-04-2024/00. 2024.04.23 Convocatoria Directorio.pdf"
)
text = _extract_text(content)

print("FECHA:", _first_match(_DATE_RE, text))
print("HORA:", _first_match(_TIME_RE, text))
print("LUGAR:", _first_match(_PLACE_RE, text))
print("MODALIDAD:", _detect_modalidad(text))
print()
print("=== ORDEN DEL DIA ===")
orden = _parse_orden_del_dia(text)
print(f"Cantidad de puntos top-level: {len(orden)}")
print(json.dumps(orden, ensure_ascii=False, indent=2))
print()
print("=== ROLES ===")
print(json.dumps(_parse_roles(text), ensure_ascii=False, indent=2))
print()
print("=== TEXTO CRUDO ALREDEDOR DE 'Señores' (repr) ===")
idx = text.lower().find("señores")
print(repr(text[max(0, idx-50):idx+600]))

"""

"""
import os, json

with open("local.settings.json", encoding="utf-8") as f:
    settings = json.load(f)
    for key, value in settings["Values"].items():
        os.environ[key] = value

from tools.people_catalog import read_catalog
catalog = read_catalog()
print(f"Total de personas cargadas: {len(catalog)}")
for p in catalog:
    print(f"  {p['nombreCompleto']} — {p['rol']}")

"""
import os, json

with open("local.settings.json", encoding="utf-8") as f:
    settings = json.load(f)
    for key, value in settings["Values"].items():
        os.environ[key] = value

import azure.functions as func
from tools.finalize_acta import register_finalize_acta_tool

app = func.FunctionApp(http_auth_level=func.AuthLevel.FUNCTION)
finalize_acta = register_finalize_acta_tool(app)

result = finalize_acta({
    "arguments": {"actaId": "78beb6bc-fc8b-475e-80ca-c95a6a29a315"}
})

result_data = json.loads(result)
print("status:", result_data.get("status"))
print("fileName:", result_data.get("fileName"))
print("webUrl:", result_data.get("webUrl"))
print("maestroActualizado:", result_data.get("maestroActualizado"))
if result_data.get("status") == "error":
    print("message:", result_data.get("message"))